# ================================================
#  backend/app.py
#  Application Flask principale
# ================================================

from flask import send_from_directory
import csv
import io
import os
from datetime import datetime, date
from functools import wraps

from flask import (
    Flask, request, jsonify, session,
    redirect, url_for, render_template_string,
    make_response, send_file
)
from flask_login import (
    LoginManager, login_user, logout_user,
    login_required, current_user
)
from flask_cors import CORS
from flask_migrate import Migrate
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from models import db, Utilisateur, Reponse

# ── Config ───────────────────────────────────
load_dotenv()

app = Flask(__name__)

import os

# ── Serve frontend static files ───────────────
FRONTEND_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'frontend')

@app.route('/')
def index():
    return send_from_directory(FRONTEND_DIR, 'index.html')

@app.route('/style.css')
def styles():
    return send_from_directory(FRONTEND_DIR, 'style.css')

@app.route('/script.js')
def scripts():
    return send_from_directory(FRONTEND_DIR, 'script.js')
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'dev-secret')
app.config['SQLALCHEMY_DATABASE_URI'] = (
    f"mysql+pymysql://{os.getenv('DB_USER')}:{os.getenv('DB_PASS')}"
    f"@{os.getenv('DB_HOST', 'localhost')}:{os.getenv('DB_PORT', '3306')}"
    f"/{os.getenv('DB_NAME')}"
)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# ── Extensions ───────────────────────────────
db.init_app(app)
migrate = Migrate(app, db)

CORS(app, supports_credentials=True, origins=[
    'http://localhost:8080',
    'http://127.0.0.1:8080',
])

login_manager = LoginManager(app)
login_manager.login_view = 'login_page'

@login_manager.user_loader
def load_user(user_id):
    return Utilisateur.query.get(int(user_id))


# ── Helpers ───────────────────────────────────
def parse_date(val):
    if not val:
        return None
    try:
        return datetime.strptime(val, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return None

def parse_int(val):
    try:
        return int(val) if val not in (None, '') else None
    except (ValueError, TypeError):
        return None

def parse_float(val):
    try:
        return float(val) if val not in (None, '') else None
    except (ValueError, TypeError):
        return None

def get_ip():
    return (request.headers.get('X-Forwarded-For', '').split(',')[0].strip()
            or request.remote_addr)

# Champs JSON à aplatir dans les exports
JSON_FIELDS = {
    'renale_type', 'diabete_ttt', 'autoimmune', 'autres_conditions',
    'atcd_familiaux', 'pe1_complic', 'pe2_complic', 'complic_ant',
}

def flatten(val):
    if isinstance(val, list):
        return ' | '.join(str(v) for v in val)
    return val or ''


# ════════════════════════════════════════════
#  API – Formulaire public
# ════════════════════════════════════════════

@app.route('/api/reponses/', methods=['POST'])
def save_reponse():
    """Enregistre une réponse du formulaire (accès public)."""
    data = request.get_json(silent=True) or request.form.to_dict()

    # Champs liste (checkboxes)
    def arr(key):
        v = data.get(key)
        if isinstance(v, list):
            return v
        if isinstance(v, str) and v:
            return [v]
        return None

    r = Reponse(
        # Section A
        id_etude        = data.get('id_etude') or None,
        date_entretien  = parse_date(data.get('date_entretien')),
        date_naissance  = parse_date(data.get('date_naissance')),
        age_actuel      = parse_int(data.get('age_actuel')),
        poids           = parse_float(data.get('poids')),
        taille          = parse_int(data.get('taille')),
        imc             = parse_float(data.get('imc')),
        imc_categorie   = data.get('imc_categorie') or None,
        education       = data.get('education') or None,
        residence       = data.get('residence') or None,
        # Section B
        ddr              = parse_date(data.get('ddr')),
        ag_semaines      = parse_int(data.get('ag_semaines')),
        ag_jours         = parse_int(data.get('ag_jours')),
        methode_datation = data.get('methode_datation') or None,
        type_grossesse   = data.get('type_grossesse') or None,
        # Section C – HTA
        hta                    = data.get('hta') or None,
        hta_date_diag          = data.get('hta_date_diag') or None,
        hta_duree              = parse_int(data.get('hta_duree')),
        hta_medicaments        = data.get('hta_medicaments') or None,
        hta_medicaments_detail = data.get('hta_medicaments_detail') or None,
        tas_avant              = parse_int(data.get('tas_avant')),
        tad_avant              = parse_int(data.get('tad_avant')),
        # Section C – Rénale
        renale            = data.get('renale') or None,
        renale_type       = arr('renale_type'),
        renale_type_autre = data.get('renale_type_autre') or None,
        dialyse           = data.get('dialyse') or None,
        # Section C – Diabète
        diabete           = data.get('diabete') or None,
        diabete_type      = data.get('diabete_type') or None,
        diabete_ttt       = arr('diabete_ttt'),
        diabete_ttt_autre = data.get('diabete_ttt_autre') or None,
        # Section C – Auto-immunes
        autoimmune        = arr('autoimmune'),
        autoimmune_autre  = data.get('autoimmune_autre') or None,
        # Section C – Autres conditions
        autres_conditions      = arr('autres_conditions'),
        cardio_detail          = data.get('cardio_detail') or None,
        thrombophilie_detail   = data.get('thrombophilie_detail') or None,
        autre_chronique_detail = data.get('autre_chronique_detail') or None,
        # Section C – Familiaux
        atcd_familiaux = arr('atcd_familiaux'),
        # Section D – Parité
        nb_grossesses    = parse_int(data.get('nb_grossesses')),
        nb_accouchements = parse_int(data.get('nb_accouchements')),
        parite           = data.get('parite') or None,
        # Section D – Prééclampsie
        atcd_pe             = data.get('atcd_pe') or None,
        pe1_annee           = parse_int(data.get('pe1_annee')),
        pe1_ag_diag         = parse_int(data.get('pe1_ag_diag')),
        pe1_severite        = data.get('pe1_severite') or None,
        pe1_ag_accouchement = parse_int(data.get('pe1_ag_accouchement')),
        pe1_complic         = arr('pe1_complic'),
        pe1_complic_autre   = data.get('pe1_complic_autre') or None,
        pe2_annee           = parse_int(data.get('pe2_annee')),
        pe2_ag_diag         = parse_int(data.get('pe2_ag_diag')),
        pe2_severite        = data.get('pe2_severite') or None,
        pe2_ag_accouchement = parse_int(data.get('pe2_ag_accouchement')),
        pe2_complic         = arr('pe2_complic'),
        pe2_complic_autre   = data.get('pe2_complic_autre') or None,
        # Section D – Autres complications
        complic_ant = arr('complic_ant'),
        # Section D – Fausses couches
        fcs        = data.get('fcs') or None,
        fcs_nombre = parse_int(data.get('fcs_nombre')),
        # Section D – Intervalle
        intervalle = data.get('intervalle') or None,
        # Métadonnées
        ip_saisie = get_ip(),
    )

    db.session.add(r)
    db.session.commit()

    return jsonify({'success': True, 'id': r.id, 'message': 'Réponse enregistrée.'}), 201


# ════════════════════════════════════════════
#  AUTH
# ════════════════════════════════════════════

@app.route('/auth/login/', methods=['GET', 'POST'])
def login_page():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))

    error = None
    username = ''

    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()

        if not username or not password:
            error = 'Veuillez remplir tous les champs.'
        else:
            user = Utilisateur.query.filter_by(username=username).first()
            if user and user.check_password(password):
                login_user(user)
                return redirect(url_for('dashboard'))
            else:
                error = 'Identifiants incorrects.'

    return render_template_string(LOGIN_HTML, error=error, username=username)


@app.route('/auth/logout/')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login_page'))


# ════════════════════════════════════════════
#  ADMIN PANEL
# ════════════════════════════════════════════

@app.route('/admin-panel/dashboard/')
@login_required
def dashboard():
    today = date.today()
    stats = {
        'total':   Reponse.query.count(),
        'today':   Reponse.query.filter(db.func.date(Reponse.cree_le) == today).count(),
        'hta_oui': Reponse.query.filter_by(hta='oui').count(),
        'pe_oui':  Reponse.query.filter_by(atcd_pe='oui').count(),
    }
    recentes = Reponse.query.order_by(Reponse.cree_le.desc()).limit(10).all()
    return render_template_string(DASHBOARD_HTML, stats=stats, recentes=recentes)


@app.route('/admin-panel/reponses/')
@login_required
def response_list():
    q    = request.args.get('q', '').strip()
    page = request.args.get('page', 1, type=int)
    qs   = Reponse.query
    if q:
        qs = qs.filter(
            db.or_(
                Reponse.id_etude.ilike(f'%{q}%'),
                db.cast(Reponse.date_entretien, db.String).ilike(f'%{q}%')
            )
        )
    pagination = qs.order_by(Reponse.cree_le.desc()).paginate(page=page, per_page=20, error_out=False)
    return render_template_string(LIST_HTML, pagination=pagination, q=q)


@app.route('/admin-panel/reponses/<int:pk>/')
@login_required
def response_detail(pk):
    r = Reponse.query.get_or_404(pk)
    return render_template_string(DETAIL_HTML, r=r)


# ════════════════════════════════════════════
#  EXPORT
# ════════════════════════════════════════════

EXPORT_FIELDS = [
    'id', 'id_etude', 'date_entretien', 'date_naissance', 'age_actuel',
    'poids', 'taille', 'imc', 'imc_categorie', 'education', 'residence',
    'ddr', 'ag_semaines', 'ag_jours', 'methode_datation', 'type_grossesse',
    'hta', 'hta_date_diag', 'hta_duree', 'hta_medicaments', 'hta_medicaments_detail',
    'tas_avant', 'tad_avant', 'renale', 'renale_type', 'renale_type_autre', 'dialyse',
    'diabete', 'diabete_type', 'diabete_ttt', 'diabete_ttt_autre',
    'autoimmune', 'autoimmune_autre', 'autres_conditions', 'cardio_detail',
    'thrombophilie_detail', 'autre_chronique_detail', 'atcd_familiaux',
    'nb_grossesses', 'nb_accouchements', 'parite', 'atcd_pe',
    'pe1_annee', 'pe1_ag_diag', 'pe1_severite', 'pe1_ag_accouchement', 'pe1_complic', 'pe1_complic_autre',
    'pe2_annee', 'pe2_ag_diag', 'pe2_severite', 'pe2_ag_accouchement', 'pe2_complic', 'pe2_complic_autre',
    'complic_ant', 'fcs', 'fcs_nombre', 'intervalle', 'ip_saisie', 'cree_le',
]


def get_export_qs():
    qs = Reponse.query
    dmin = request.args.get('date_min')
    dmax = request.args.get('date_max')
    if dmin:
        qs = qs.filter(Reponse.date_entretien >= dmin)
    if dmax:
        qs = qs.filter(Reponse.date_entretien <= dmax)
    return qs.all()


@app.route('/api/export/csv/')
@login_required
def export_csv():
    filename = f"reponses_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    output   = io.StringIO()
    output.write('\ufeff')  # BOM UTF-8 pour Excel
    writer = csv.writer(output, delimiter=';')
    writer.writerow(EXPORT_FIELDS)

    for row in get_export_qs():
        line = []
        for f in EXPORT_FIELDS:
            val = getattr(row, f, '')
            line.append(flatten(val) if f in JSON_FIELDS else (val if val is not None else ''))
        writer.writerow(line)

    resp = make_response(output.getvalue())
    resp.headers['Content-Type']        = 'text/csv; charset=utf-8'
    resp.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


@app.route('/api/export/excel/')
@login_required
def export_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Réponses'

    hdr_font  = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill  = PatternFill('solid', fgColor='1A2640')
    alt_fill  = PatternFill('solid', fgColor='FDF4F6')
    center    = Alignment(horizontal='center', vertical='center')

    for ci, field in enumerate(EXPORT_FIELDS, 1):
        cell = ws.cell(row=1, column=ci, value=field.upper())
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center

    for ri, row in enumerate(get_export_qs(), 2):
        for ci, f in enumerate(EXPORT_FIELDS, 1):
            val  = getattr(row, f, '')
            val  = flatten(val) if f in JSON_FIELDS else (str(val) if val is not None else '')
            cell = ws.cell(row=ri, column=ci, value=val)
            if ri % 2 == 0:
                cell.fill = alt_fill

    for col in ws.columns:
        w = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(w + 4, 40)

    buf      = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"reponses_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ════════════════════════════════════════════
#  TEMPLATES INLINE
# ════════════════════════════════════════════

_BASE_STYLE = """
<link href="https://fonts.googleapis.com/css2?family=Playfair+Display:wght@400;600&family=DM+Sans:wght@300;400;500&display=swap" rel="stylesheet">
<style>
:root{--rose:#c8496a;--rose-light:#f5dce3;--rose-pale:#fdf4f6;--navy:#1a2640;--navy-mid:#2e3f5c;--text:#2c2c2c;--muted:#6b7280;--border:#dde3ec;--shadow:0 2px 16px rgba(26,38,64,.07);--radius:10px}
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'DM Sans',sans-serif;background:var(--rose-pale);color:var(--text)}
.topbar{background:var(--navy);color:white;padding:14px 28px;display:flex;align-items:center;justify-content:space-between}
.topbar h1{font-family:'Playfair Display',serif;font-size:1.1rem}
.topbar nav a{color:rgba(255,255,255,.75);text-decoration:none;font-size:13px;margin-left:16px}
.topbar nav a:hover,.topbar nav a.active{color:#f9b8c8}
.wrap{max-width:1100px;margin:0 auto;padding:32px 20px}
.stats{display:grid;grid-template-columns:repeat(auto-fill,minmax(200px,1fr));gap:16px;margin-bottom:28px}
.stat{background:white;border-radius:12px;padding:22px 24px;border:1px solid var(--border)}
.stat .val{font-family:'Playfair Display',serif;font-size:2.2rem;color:var(--navy);font-weight:600}
.stat .lbl{font-size:13px;color:var(--muted);margin-top:4px}
.stat.accent .val{color:var(--rose)}
.card{background:white;border-radius:12px;border:1px solid var(--border);box-shadow:var(--shadow);overflow:hidden;margin-bottom:20px}
.card h2{font-family:'Playfair Display',serif;font-size:1rem;padding:18px 24px;border-bottom:1px solid var(--border);color:var(--navy)}
table{width:100%;border-collapse:collapse;font-size:13px}
thead th{background:var(--navy);color:white;padding:10px 14px;text-align:left;font-weight:500}
tbody tr:nth-child(even){background:var(--rose-pale)}
tbody td{padding:9px 14px}
.badge{display:inline-block;padding:2px 10px;border-radius:99px;font-size:11px;font-weight:500}
.badge-oui{background:#d4edda;color:#155724}
.badge-non{background:#f8d7da;color:#721c24}
.badge-nsp{background:#fff3cd;color:#856404}
.btn{background:var(--navy);color:white;border:none;padding:9px 18px;border-radius:8px;font-family:'DM Sans',sans-serif;font-size:13px;cursor:pointer;text-decoration:none;display:inline-block}
.btn:hover{background:var(--rose)}
.btn.green{background:#1e7e34}.btn.green:hover{background:#155724}
.btn-link{color:var(--rose);text-decoration:none;font-size:12px;font-weight:500}
.export-bar{background:white;border-radius:12px;padding:18px 24px;border:1px solid var(--border);margin-bottom:28px;display:flex;align-items:center;gap:16px;flex-wrap:wrap}
.toolbar{display:flex;align-items:center;gap:12px;margin-bottom:20px}
.toolbar input{max-width:280px;padding:9px 14px;border:1.5px solid var(--border);border-radius:var(--radius);font-family:'DM Sans',sans-serif;font-size:14px;outline:none}
.toolbar input:focus{border-color:var(--rose)}
input[type=date]{padding:9px 14px;border:1.5px solid var(--border);border-radius:var(--radius);font-family:'DM Sans',sans-serif;font-size:13px;outline:none}
.pagination{display:flex;gap:6px;margin-top:18px;justify-content:center}
.pagination a,.pagination span{padding:7px 13px;border-radius:7px;font-size:13px;text-decoration:none;border:1px solid var(--border);color:var(--navy)}
.pagination a:hover{background:var(--rose-pale);border-color:var(--rose)}
.pagination .current{background:var(--rose);color:white;border-color:var(--rose)}
.detail-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(230px,1fr));gap:14px}
.di label{display:block;font-size:11px;text-transform:uppercase;letter-spacing:.8px;color:var(--rose);margin-bottom:4px;font-weight:500}
.di .val{font-size:14px}
.sep{border-top:1px solid var(--border);margin:20px 0}
.sec-title{font-size:13px;font-weight:500;text-transform:uppercase;letter-spacing:1px;color:var(--rose);margin-bottom:16px;padding-bottom:8px;border-bottom:1px solid var(--rose-light)}
.back{display:inline-flex;align-items:center;gap:6px;color:var(--navy);text-decoration:none;font-size:13px;font-weight:500;margin-bottom:20px}
.back:hover{color:var(--rose)}
.dheader{background:linear-gradient(90deg,var(--navy),var(--navy-mid));color:white;padding:16px 24px;display:flex;justify-content:space-between;align-items:center}
.dheader h2{font-family:'Playfair Display',serif;font-size:1rem}
.dheader span{font-size:12px;opacity:.75}
.dbody{padding:24px}
</style>
"""

_TOPBAR = """
<div class="topbar">
  <h1>🩺 Étude Prééclampsie</h1>
  <nav>
    <a href="/admin-panel/dashboard/" class="{{ 'active' if active=='dashboard' else '' }}">Dashboard</a>
    <a href="/admin-panel/reponses/"  class="{{ 'active' if active=='list' else '' }}">Réponses</a>
    <a href="http://localhost:8080" target="_blank">Formulaire ↗</a>
    <span style="color:rgba(255,255,255,.3);margin-left:16px">|</span>
    <a href="/auth/logout/" style="margin-left:8px">Déconnexion</a>
  </nav>
</div>
"""

def badge(val):
    if val == 'oui':   return '<span class="badge badge-oui">oui</span>'
    if val == 'non':   return '<span class="badge badge-non">non</span>'
    if val:            return f'<span class="badge badge-nsp">{val}</span>'
    return '–'

app.jinja_env.globals['badge'] = badge

LOGIN_HTML = """<!DOCTYPE html><html lang="fr"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Connexion</title>""" + _BASE_STYLE + """
<style>
body{display:flex;align-items:center;justify-content:center;min-height:100vh}
.card{background:white;border-radius:16px;box-shadow:0 4px 32px rgba(26,38,64,.12);padding:48px 40px;width:100%;max-width:420px;border:1px solid var(--border)}
.card h1{font-family:'Playfair Display',serif;font-size:1.6rem;color:var(--navy);margin-bottom:6px}
.card p{color:var(--muted);font-size:14px;margin-bottom:32px}
.err{background:#fff0f3;border:1px solid #fbb6c8;color:#a33057;border-radius:8px;padding:10px 14px;font-size:14px;margin-bottom:20px}
.field{margin-bottom:18px}
.field label{display:block;font-size:13px;font-weight:500;color:var(--navy-mid);margin-bottom:6px}
.field input{width:100%;padding:10px 14px;border:1.5px solid var(--border);border-radius:var(--radius);font-family:'DM Sans',sans-serif;font-size:14px;outline:none}
.field input:focus{border-color:var(--rose)}
.btn-full{width:100%;background:linear-gradient(135deg,var(--rose),#a33057);color:white;border:none;padding:14px;font-family:'DM Sans',sans-serif;font-size:15px;font-weight:500;border-radius:50px;cursor:pointer}
</style></head><body>
<div class="card">
  <h1>Connexion</h1>
  <p>Accès réservé à l'équipe de recherche</p>
  {% if error %}<div class="err">{{ error }}</div>{% endif %}
  <form method="POST">
    <div class="field"><label>Nom d'utilisateur</label>
      <input type="text" name="username" value="{{ username }}" required autofocus></div>
    <div class="field"><label>Mot de passe</label>
      <input type="password" name="password" required></div>
    <button type="submit" class="btn-full">Se connecter</button>
  </form>
</div></body></html>"""

DASHBOARD_HTML = """<!DOCTYPE html><html lang="fr"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Dashboard</title>""" + _BASE_STYLE + "</head><body>" + _TOPBAR + """
<div class="wrap">
  <div class="stats">
    <div class="stat"><div class="val">{{ stats.total }}</div><div class="lbl">Questionnaires total</div></div>
    <div class="stat accent"><div class="val">{{ stats.today }}</div><div class="lbl">Saisies aujourd'hui</div></div>
    <div class="stat"><div class="val">{{ stats.hta_oui }}</div><div class="lbl">Patientes avec HTA</div></div>
    <div class="stat"><div class="val">{{ stats.pe_oui }}</div><div class="lbl">Antécédents de PE</div></div>
  </div>
  <div class="export-bar">
    <strong style="font-size:14px;color:var(--navy)">Exporter :</strong>
    <div style="display:flex;align-items:center;gap:8px">
      <label style="font-size:13px">Du</label><input type="date" id="dmin">
      <label style="font-size:13px">Au</label><input type="date" id="dmax">
    </div>
    <a class="btn green" id="bcsv" href="#">⬇ CSV</a>
    <a class="btn"       id="bxls" href="#">⬇ Excel</a>
  </div>
  <div class="card">
    <h2>10 dernières saisies</h2>
    <table><thead><tr><th>#</th><th>ID Étude</th><th>Date</th><th>Âge</th><th>HTA</th><th>Atcd PE</th><th>Enregistré le</th><th></th></tr></thead>
    <tbody>
    {% for r in recentes %}
    <tr>
      <td>{{ r.id }}</td><td>{{ r.id_etude or '–' }}</td>
      <td>{{ r.date_entretien or '–' }}</td>
      <td>{{ (r.age_actuel|string + ' ans') if r.age_actuel else '–' }}</td>
      <td>{{ badge(r.hta) }}</td><td>{{ badge(r.atcd_pe) }}</td>
      <td>{{ r.cree_le.strftime('%d/%m/%Y %H:%M') }}</td>
      <td><a class="btn-link" href="/admin-panel/reponses/{{ r.id }}/">Voir →</a></td>
    </tr>
    {% else %}<tr><td colspan="8" style="text-align:center;color:var(--muted);padding:24px">Aucune réponse.</td></tr>
    {% endfor %}
    </tbody></table>
  </div>
</div>
<script>
function url(fmt){
  const mn=document.getElementById('dmin').value, mx=document.getElementById('dmax').value;
  let u='/api/export/'+fmt+'/'; const p=[];
  if(mn) p.push('date_min='+mn); if(mx) p.push('date_max='+mx);
  return p.length ? u+'?'+p.join('&') : u;
}
document.getElementById('bcsv').onclick=e=>{e.preventDefault();window.location.href=url('csv')};
document.getElementById('bxls').onclick=e=>{e.preventDefault();window.location.href=url('excel')};
</script>
</body></html>"""

LIST_HTML = """<!DOCTYPE html><html lang="fr"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Réponses</title>""" + _BASE_STYLE + "</head><body>" + _TOPBAR + """
<div class="wrap">
  <div class="toolbar">
    <form method="GET" style="display:flex;gap:8px;flex:1">
      <input type="text" name="q" value="{{ q }}" placeholder="Rechercher par ID ou date…">
      <button type="submit" class="btn">Rechercher</button>
      {% if q %}<a href="/admin-panel/reponses/" class="btn" style="background:var(--muted)">✕</a>{% endif %}
    </form>
    <span style="font-size:13px;color:var(--muted)">{{ pagination.total }} résultat{{ 's' if pagination.total > 1 else '' }}</span>
  </div>
  <div class="card">
    <table><thead><tr><th>#</th><th>ID Étude</th><th>Date</th><th>Âge</th><th>Résidence</th><th>Parité</th><th>HTA</th><th>Atcd PE</th><th>Enregistré le</th><th></th></tr></thead>
    <tbody>
    {% for r in pagination.items %}
    <tr>
      <td>{{ r.id }}</td><td>{{ r.id_etude or '–' }}</td>
      <td>{{ r.date_entretien or '–' }}</td>
      <td>{{ (r.age_actuel|string+' ans') if r.age_actuel else '–' }}</td>
      <td>{{ r.residence or '–' }}</td><td>{{ r.parite or '–' }}</td>
      <td>{{ badge(r.hta) }}</td><td>{{ badge(r.atcd_pe) }}</td>
      <td>{{ r.cree_le.strftime('%d/%m/%Y %H:%M') }}</td>
      <td><a class="btn-link" href="/admin-panel/reponses/{{ r.id }}/">Voir →</a></td>
    </tr>
    {% else %}<tr><td colspan="10" style="text-align:center;color:var(--muted);padding:24px">Aucune réponse.</td></tr>
    {% endfor %}
    </tbody></table>
  </div>
  <div class="pagination">
    {% if pagination.has_prev %}<a href="?page={{ pagination.prev_num }}&q={{ q }}">‹</a>{% endif %}
    {% for p in pagination.iter_pages(left_edge=1,right_edge=1,left_current=2,right_current=2) %}
      {% if p %}
        {% if p == pagination.page %}<span class="current">{{ p }}</span>
        {% else %}<a href="?page={{ p }}&q={{ q }}">{{ p }}</a>{% endif %}
      {% else %}<span>…</span>{% endif %}
    {% endfor %}
    {% if pagination.has_next %}<a href="?page={{ pagination.next_num }}&q={{ q }}">›</a>{% endif %}
  </div>
</div></body></html>"""

DETAIL_HTML = """<!DOCTYPE html><html lang="fr"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Réponse #{{ r.id }}</title>""" + _BASE_STYLE + "</head><body>" + _TOPBAR + """
<div class="wrap">
  <a href="/admin-panel/reponses/" class="back">← Retour à la liste</a>
  <div class="card">
    <div class="dheader"><h2>Réponse #{{ r.id }} — {{ r.id_etude or 'N/A' }}</h2>
      <span>{{ r.cree_le.strftime('Enregistré le %d/%m/%Y à %H:%M') }}</span></div>
    <div class="dbody">
      <div class="sec-title">Section A – Informations démographiques</div>
      <div class="detail-grid">
        <div class="di"><label>Date entretien</label><div class="val">{{ r.date_entretien or '–' }}</div></div>
        <div class="di"><label>Date naissance</label><div class="val">{{ r.date_naissance or '–' }}</div></div>
        <div class="di"><label>Âge</label><div class="val">{{ (r.age_actuel|string+' ans') if r.age_actuel else '–' }}</div></div>
        <div class="di"><label>Poids</label><div class="val">{{ (r.poids|string+' kg') if r.poids else '–' }}</div></div>
        <div class="di"><label>Taille</label><div class="val">{{ (r.taille|string+' cm') if r.taille else '–' }}</div></div>
        <div class="di"><label>IMC</label><div class="val">{{ (r.imc|string+' kg/m² ('+r.imc_categorie+')') if r.imc else '–' }}</div></div>
        <div class="di"><label>Éducation</label><div class="val">{{ r.education or '–' }}</div></div>
        <div class="di"><label>Résidence</label><div class="val">{{ r.residence or '–' }}</div></div>
      </div>
      <div class="sep"></div>
      <div class="sec-title">Section B – Grossesse actuelle</div>
      <div class="detail-grid">
        <div class="di"><label>DDR</label><div class="val">{{ r.ddr or '–' }}</div></div>
        <div class="di"><label>Âge gestationnel</label><div class="val">{% if r.ag_semaines is not none %}{{ r.ag_semaines }}SA + {{ r.ag_jours }}j{% else %}–{% endif %}</div></div>
        <div class="di"><label>Méthode datation</label><div class="val">{{ r.methode_datation or '–' }}</div></div>
        <div class="di"><label>Type grossesse</label><div class="val">{{ r.type_grossesse or '–' }}</div></div>
      </div>
      <div class="sep"></div>
      <div class="sec-title">Section C – Antécédents médicaux</div>
      <div class="detail-grid">
        <div class="di"><label>HTA</label><div class="val">{{ badge(r.hta) }}</div></div>
        <div class="di"><label>Maladie rénale</label><div class="val">{{ badge(r.renale) }}</div></div>
        <div class="di"><label>Diabète</label><div class="val">{{ badge(r.diabete) }}</div></div>
        <div class="di"><label>Auto-immunes</label><div class="val">{{ (r.autoimmune | join(', ')) if r.autoimmune else '–' }}</div></div>
        <div class="di"><label>Autres conditions</label><div class="val">{{ (r.autres_conditions | join(', ')) if r.autres_conditions else '–' }}</div></div>
        <div class="di"><label>Antécédents familiaux</label><div class="val">{{ (r.atcd_familiaux | join(', ')) if r.atcd_familiaux else '–' }}</div></div>
      </div>
      <div class="sep"></div>
      <div class="sec-title">Section D – Antécédents de grossesse</div>
      <div class="detail-grid">
        <div class="di"><label>Nb grossesses</label><div class="val">{{ r.nb_grossesses or '–' }}</div></div>
        <div class="di"><label>Parité</label><div class="val">{{ r.parite or '–' }}</div></div>
        <div class="di"><label>Atcd prééclampsie</label><div class="val">{{ badge(r.atcd_pe) }}</div></div>
        <div class="di"><label>Complications</label><div class="val">{{ (r.complic_ant | join(', ')) if r.complic_ant else '–' }}</div></div>
        <div class="di"><label>Fausses couches</label><div class="val">{% if r.fcs == 'oui' %}Oui ({{ r.fcs_nombre }}){% else %}{{ r.fcs or '–' }}{% endif %}</div></div>
        <div class="di"><label>Intervalle</label><div class="val">{{ r.intervalle or '–' }}</div></div>
      </div>
    </div>
  </div>
</div></body></html>"""


# ════════════════════════════════════════════
#  Initialisation BDD + compte admin
# ════════════════════════════════════════════

def init_db():
    with app.app_context():
        db.create_all()
        if not Utilisateur.query.filter_by(username='admin').first():
            admin = Utilisateur(
                nom='Administrateur',
                username='admin',
                email='admin@etude.local',
                role='admin'
            )
            admin.set_password('Admin1234!')
            db.session.add(admin)
            db.session.commit()
            print('✅ Compte admin créé : admin / Admin1234!')


if __name__ == '__main__':
    init_db()
    app.run(host='0.0.0.0', port=8000, debug=True)
